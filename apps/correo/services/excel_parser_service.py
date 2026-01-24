"""
Servicio para procesar archivos Excel con información de estudiantes.
"""
import openpyxl
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import re


class ExcelParserService:
    """
    Servicio para parsear archivos Excel con datos de estudiantes.
    """
    
    @staticmethod
    def parse_excel(file_path):
        """
        Parsea un archivo Excel y extrae información de estudiantes.
        
        Args:
            file_path: Ruta al archivo Excel
            
        Returns:
            dict: {
                'success': bool,
                'data': list de dicts con {full_name, email, certificate_link},
                'errors': list de mensajes de error,
                'warnings': list de advertencias (duplicados, etc)
            }
        """
        result = {
            'success': False,
            'data': [],
            'errors': [],
            'warnings': []
        }
        
        try:
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Detectar las columnas
            headers = {}
            first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            for idx, header in enumerate(first_row):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'nombre' in header_lower:
                        headers['name'] = idx
                    elif 'correo' in header_lower or 'email' in header_lower:
                        headers['email'] = idx
                    elif 'link' in header_lower or 'certificado' in header_lower:
                        headers['link'] = idx
            
            # Validar que existan las columnas necesarias
            required_columns = ['name', 'email', 'link']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                result['errors'].append(
                    f"El archivo Excel debe contener columnas: 'Nombre', 'Correo/Email', 'Link/Certificado'. "
                    f"Faltan: {', '.join(missing_columns)}"
                )
                return result
            
            # Procesar las filas
            seen_emails = set()
            row_number = 1
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_number += 1
                
                # Obtener valores
                full_name = row[headers['name']] if len(row) > headers['name'] else None
                email = row[headers['email']] if len(row) > headers['email'] else None
                certificate_link = row[headers['link']] if len(row) > headers['link'] else None
                
                # Saltar filas vacías
                if not full_name and not email and not certificate_link:
                    continue
                
                # Validar que todos los campos tengan valor
                if not full_name or not email or not certificate_link:
                    result['warnings'].append(
                        f"Fila {row_number}: campos incompletos (se omitirá esta fila)"
                    )
                    continue
                
                # Limpiar y validar nombre
                full_name = str(full_name).strip()
                if len(full_name) < 3:
                    result['warnings'].append(
                        f"Fila {row_number}: nombre muy corto '{full_name}' (se omitirá)"
                    )
                    continue
                
                # Limpiar y validar email
                email = str(email).strip().lower()
                try:
                    validate_email(email)
                except ValidationError:
                    result['warnings'].append(
                        f"Fila {row_number}: correo inválido '{email}' (se omitirá)"
                    )
                    continue
                
                # Verificar duplicados
                if email in seen_emails:
                    result['warnings'].append(
                        f"Fila {row_number}: correo duplicado '{email}' (se omitirá)"
                    )
                    continue
                
                seen_emails.add(email)
                
                # Limpiar y validar link
                certificate_link = str(certificate_link).strip()
                if not ExcelParserService._is_valid_url(certificate_link):
                    result['warnings'].append(
                        f"Fila {row_number}: URL inválida '{certificate_link}' (se incluirá de todos modos)"
                    )
                
                # Agregar a los datos
                result['data'].append({
                    'full_name': full_name,
                    'email': email,
                    'certificate_link': certificate_link
                })
            
            if not result['data']:
                result['errors'].append(
                    "No se encontraron datos válidos en el archivo Excel"
                )
                return result
            
            result['success'] = True
            
        except Exception as e:
            result['errors'].append(f"Error al procesar el archivo Excel: {str(e)}")
        
        return result
    
    @staticmethod
    def _is_valid_url(url):
        """
        Valida si una cadena es una URL válida.
        """
        url_pattern = re.compile(
            r'^https?://'  # http:// o https://
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'  # dominio
            r'localhost|'  # localhost
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # IP
            r'(?::\d+)?'  # puerto opcional
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        
        return url_pattern.match(url) is not None
