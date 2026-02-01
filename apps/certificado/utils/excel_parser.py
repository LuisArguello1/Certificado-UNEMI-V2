"""
Parser de archivos Excel con estudiantes.

Extrae y valida los datos de estudiantes desde archivos Excel.
"""

import openpyxl
import re
import logging
from typing import List, Dict
from django.core.exceptions import ValidationError


logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """
    Error personalizado para errores de parsing de Excel.
    """
    pass


class ExcelParser:
    """
    Clase para parsear archivos Excel con estudiantes.
    
    Formato esperado:
        - Columna 1: NOMBRES COMPLETOS
        - Columna 2: CORREO ELECTRONICO
    
    Retorna lista de diccionarios con los datos validados.
    """
    
    # Nombres alternativos permitidos para headers (case-insensitive, sin tildes)
    NOMBRES_HEADERS = [
        'NOMBRES COMPLETOS',
        'NOMBRE COMPLETO',
        'NOMBRES',
        'NOMBRE',
        'ESTUDIANTE',
        'ESTUDIANTES',
        'PARTICIPANTE',
        'PARTICIPANTES'
    ]
    
    CORREO_HEADERS = [
        'CORREO ELECTRONICO',
        'CORREO ELECTRÓNICO',
        'CORREOS ELECTRONICOS',
        'CORREOS ELECTRÓNICOS',
        'CORREO',
        'CORREOS',
        'EMAIL',
        'EMAILS',
        'E-MAIL',
        'E-MAILS',
        'MAIL',
        'MAILS'
    ]
    
    @staticmethod
    def normalize_text(text):
        """
        Normaliza texto para comparación: quita tildes, convierte a mayúsculas, quita espacios extras.
        """
        if not text:
            return ""
        import unicodedata
        # Quitar tildes
        text = ''.join(
            c for c in unicodedata.normalize('NFD', str(text))
            if unicodedata.category(c) != 'Mn'
        )
        # Mayúsculas y quitar espacios extras
        return ' '.join(text.upper().split())
    
    @staticmethod
    def sanitize_value(value):
        """
        Limpia caracteres invisibles y espacios extras.
        """
        if value is None:
            return ""
        
        # Convertir a string
        text = str(value)
        
        # Eliminar caracteres invisibles específicos (zero-width space, etc.)
        # \u200b: Zero width space
        # \ufeff: Byte order mark
        # \u200c: Zero width non-joiner
        # \u200d: Zero width joiner
        invisible_chars = ['\u200b', '\ufeff', '\u200c', '\u200d']
        for char in invisible_chars:
            text = text.replace(char, '')
            
        return text.strip()
    
    def __init__(self, file_path_or_file):
        """
        Inicializa el parser con un archivo Excel.
        
        Args:
            file_path_or_file: Ruta al archivo  o File object de Django
        """
        self.file = file_path_or_file
        self.workbook = None
        self.worksheet = None
        self.nombres_col_index = None
        self.correo_col_index = None
    
    def parse(self) -> List[Dict[str, str]]:
        """
        Parsea el archivo Excel y retorna lista de estudiantes.
        
        Returns:
            Lista de diccionarios:
            [
                {"nombres_completos": "Juan Pérez", "correo_electronico": "juan@example.com"},
                ...
            ]
        
        Raises:
            ExcelParseError: Si el formato es inválido o hay errores
        """
        try:
            # Cargar workbook
            self.workbook = openpyxl.load_workbook(self.file, read_only=True, data_only=True)
            self.worksheet = self.workbook.active
            
            # Identificar columnas
            self._identify_columns()
            
            # Extraer datos
            estudiantes = self._extract_data()
            
            # Validar datos
            self._validate_data(estudiantes)
            
            logger.info(f"Excel parseado exitosamente: {len(estudiantes)} estudiante(s) encontrado(s)")
            
            return estudiantes
            
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ExcelParseError("El archivo no es un Excel válido (.xlsx, .xls)")
        except Exception as e:
            logger.error(f"Error al parsear Excel: {str(e)}")
            raise ExcelParseError(f"Error al procesar el archivo: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _identify_columns(self):
        """
        Identifica las columnas de nombres y correo buscando en las primeras filas.
        
        Scanea hasta 10 filas buscando los headers requeridos.
        
        Raises:
            ExcelParseError: Si no se encuentran las columnas requeridas
        """
        if not self.worksheet or self.worksheet.max_row < 1:
            raise ExcelParseError("El archivo Excel está vacío")
        
        # Escanear primeras 10 filas o hasta el final
        max_scan_rows = min(10, self.worksheet.max_row)
        header_row_index = None
        
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, max_row=max_scan_rows), start=1):
            # Obtener valores normalizados de la fila
            row_values = []
            for cell in row:
                value = self.sanitize_value(cell.value)
                row_values.append(self.normalize_text(value))
            
            # Buscar si esta fila tiene AMBOS headers candidatos
            has_nombre = any(any(n in h for n in [self.normalize_text(x) for x in self.NOMBRES_HEADERS]) for h in row_values)
            has_correo = any(any(c in h for c in [self.normalize_text(x) for x in self.CORREO_HEADERS]) for h in row_values)
            
            if has_nombre and has_correo:
                header_row_index = row_idx
                logger.info(f"Headers encontrados en fila {row_idx}")
                break
        
        if header_row_index is None:
            raise ExcelParseError(
                "No se encontraron los encabezados 'NOMBRES' y 'CORREO' en las primeras 10 filas. "
                "Asegúrese de que el archivo tenga estas columnas."
            )
            
        # Procesar la fila encontrada para obtener índices exactos
        self.header_row = header_row_index
        headers = []
        # Leer fila de headers real
        for row in self.worksheet.iter_rows(min_row=header_row_index, max_row=header_row_index):
            headers = [self.sanitize_value(cell.value) for cell in row]
            
        normalized_headers = [self.normalize_text(h) for h in headers]
        
        # Buscar índice de Nombres
        for idx, norm_header in enumerate(normalized_headers):
            for nombre_option in self.NOMBRES_HEADERS:
                if self.normalize_text(nombre_option) == norm_header or  self.normalize_text(nombre_option) in norm_header:
                     self.nombres_col_index = idx
                     break
            if self.nombres_col_index is not None:
                break
                
        # Buscar índice de Correo
        for idx, norm_header in enumerate(normalized_headers):
            for correo_option in self.CORREO_HEADERS:
                if self.normalize_text(correo_option) == norm_header or self.normalize_text(correo_option) in norm_header:
                    self.correo_col_index = idx
                    break
            if self.correo_col_index is not None:
                break
        
        # Validación final de índices
        if self.nombres_col_index is None or self.correo_col_index is None:
             raise ExcelParseError("Error interno identificando columnas después de detectar headers.")
    
    def _extract_data(self) -> List[Dict[str, str]]:
        """
        Extrae los datos de las columnas identificadas.
        
        Returns:
            Lista de diccionarios con datos crudos
        """
        estudiantes = []
        
        # Iterar desde la fila siguiente al header
        start_row = self.header_row + 1
        
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=start_row), start=start_row):
            # Asegurar que la fila tenga suficientes celdas
            if len(row) <= max(self.nombres_col_index, self.correo_col_index):
                continue
                
            nombres_cell = row[self.nombres_col_index]
            correo_cell = row[self.correo_col_index]
            
            nombres = self.sanitize_value(nombres_cell.value)
            correo = self.sanitize_value(correo_cell.value)
            
            # Saltar filas completamente vacías
            if not nombres and not correo:
                continue

            estudiantes.append({
                'nombres_completos': nombres,
                'correo_electronico': correo,
                'row_number': row_idx
            })
        
        if not estudiantes:
            raise ExcelParseError(
                "No se encontraron estudiantes en el archivo. "
                "Asegúrese de que hay datos después de la fila de encabezados."
            )
        
        return estudiantes
    
    def _validate_data(self, estudiantes: List[Dict[str, str]]):
        """
        Valida los datos extraídos.
        """
        errores = []
        correos_vistos = set()
        nombres_vistos = set()
        
        # Regex para validar email
        # Permitimos '+' para alias y dominios de alto nivel largos
        # Estructura: [usuario] @ [dominio] . [tld]
        # usuario: alfanumérico, ., _, -, +
        email_regex = re.compile(r'^[a-zA-Z0-9.+_-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        
        for estudiante in estudiantes:
            row_num = estudiante['row_number']
            nombres = estudiante['nombres_completos']
            correo = estudiante['correo_electronico']
            
            # --- VALIDACIÓN DE NOMBRES ---
            
            # Validar longitud
            if len(nombres) > 300:
                errores.append(f"Fila {row_num}: Nombre demasiado largo ({len(nombres)} caracteres).")
            
            # Validar que no esté vacío
            if not nombres or nombres.isspace():
                errores.append(f"Fila {row_num}: El campo 'Nombres' está vacío.")
            else:
                # Validar Nombres Duplicados
                nombres_norm = self.normalize_text(nombres)
                if nombres_norm in nombres_vistos:
                   errores.append(f"Fila {row_num}: Nombre duplicado: '{nombres}' ya existe en el archivo.")
                else:
                    nombres_vistos.add(nombres_norm)
            
            # --- VALIDACIÓN DE CORREOS ---
            
            # Validar que no esté vacío
            if not correo or correo.isspace():
                errores.append(f"Fila {row_num}: El campo 'Correo Electrónico' está vacío.")
                continue # No seguir validando este correo
                
            correo_stripped = correo.strip()
            
            # Validar formato con Regex estricto
            if not email_regex.match(correo_stripped):
                errores.append(
                    f"Fila {row_num}: Formato de correo inválido: '{correo}'. "
                    f"No se permiten tildes, espacios ni caracteres especiales. Solo letras, números, '.', '_', '-'."
                )
                continue
            
            # Validar duplicados de correo
            correo_lower = correo_stripped.lower()
            if correo_lower in correos_vistos:
                errores.append(f"Fila {row_num}: Correo duplicado: '{correo}' ya existe en el archivo.")
            else:
                correos_vistos.add(correo_lower)
            
            # Validaciones extra de estructura
            local_part = correo_stripped.split('@')[0]
            if local_part.startswith('.') or local_part.startswith('-') or local_part.endswith('.') or local_part.endswith('-'):
                errores.append(f"Fila {row_num}: El correo no puede empezar o terminar con punto o guión.")
            
            if '..' in correo_stripped:
                errores.append(f"Fila {row_num}: El correo no puede tener puntos consecutivos.")

        # Si hay errores, lanzar excepción con todos los detalles
        if errores:
            # Limitar número de errores mostrados para no saturar la pantalla
            if len(errores) > 20:
                 errores_mostrados = errores[:20]
                 errores_mostrados.append(f"... y {len(errores) - 20} errores más.")
                 mensaje_full = "\n".join(errores_mostrados)
            else:
                 mensaje_full = "\n".join(errores)
                
            mensaje_error = f"Se encontraron errores en el archivo Excel:\n\n{mensaje_full}"
            logger.warning(mensaje_error) # Changed from error to warning
            raise ExcelParseError(mensaje_error)


def parse_excel_estudiantes(file) -> List[Dict[str, str]]:
    """
    Función helper para parsear Excel de estudiantes.
    
    Args:
        file: Archivo de Django (UploadedFile)
    
    Returns:
        Lista de diccionarios con estudiantes validados
    
    Raises:
        ExcelParseError: Si hay errores de parsing o validación
    
    Ejemplo:
        >>> estudiantes = parse_excel_estudiantes(request.FILES['archivo_excel'])
        >>> for est in estudiantes:
        >>>     print(est['nombres_completos'], est['correo_electronico'])
    """
    parser = ExcelParser(file)
    return parser.parse()
